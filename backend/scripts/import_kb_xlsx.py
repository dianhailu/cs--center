#!/usr/bin/env python3
"""Import PinGo CS Knowledge Base xlsx → knowledge/kb_v2.json + merge faq.json.

Usage:
  python scripts/import_kb_xlsx.py /path/to/PinGo_Customer_Service_Knowledge_Base_V2.0.xlsx
  python scripts/import_kb_xlsx.py  # uses default Downloads path if present
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

BACKEND = Path(__file__).resolve().parents[1]
KNOWLEDGE = BACKEND / "knowledge"
DEFAULT_XLSX = Path.home() / "Downloads" / "PinGo_Customer_Service_Knowledge_Base_V2.0.xlsx"

FAQ_SHEETS = {
    "02-产品介绍",
    "03-注册OTP",
    "04-KYC",
    "05-授信额度",
    "06-借款",
    "07-还款",
    "08-逾期",
    "09-账户安全",
    "10-投诉处理",
}

CAT_MAP = {
    "产品": {"id": "Produk", "en": "Product", "zh": "产品"},
    "注册": {"id": "Pendaftaran", "en": "Registration", "zh": "注册"},
    "KYC": {"id": "KYC", "en": "KYC", "zh": "KYC"},
    "额度": {"id": "Limit", "en": "Limit", "zh": "额度"},
    "借款": {"id": "Pinjaman", "en": "Loan", "zh": "借款"},
    "还款": {"id": "Pembayaran", "en": "Repayment", "zh": "还款"},
    "逾期": {"id": "Keterlambatan", "en": "Overdue", "zh": "逾期"},
    "安全": {"id": "Keamanan", "en": "Security", "zh": "安全"},
    "投诉": {"id": "Komplain", "en": "Complaint", "zh": "投诉"},
}


def _norm(s: str) -> str:
    s = (s or "").lower().strip()
    s = re.sub(r"^[\d\.\s]+", "", s)
    s = re.sub(r"[^\w\u4e00-\u9fff]+", "", s, flags=re.UNICODE)
    return s


def _cell(row: tuple, i: int) -> str:
    if i >= len(row) or row[i] is None:
        return ""
    return str(row[i]).strip()


def load_xlsx(path: Path) -> tuple[list[dict], list[dict], list[dict]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl required: pip install openpyxl") from exc

    wb = openpyxl.load_workbook(path, data_only=True)
    faq_items: list[dict] = []
    escalation: list[dict] = []
    qa_standards: list[dict] = []

    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if name in FAQ_SHEETS:
            for row in rows[1:]:
                if not row:
                    continue
                cat = _cell(row, 0)
                q_zh = _cell(row, 1)
                q_id = _cell(row, 2)
                a_zh = _cell(row, 3)
                a_id = _cell(row, 4)
                note = _cell(row, 5)
                if not (q_zh or q_id) or not (a_zh or a_id):
                    continue
                cat_obj = CAT_MAP.get(cat) or {
                    "id": cat or "Umum",
                    "en": cat or "General",
                    "zh": cat or "通用",
                }
                faq_items.append(
                    {
                        "source": "kb_v2_xlsx",
                        "sheet": name,
                        "note": note or None,
                        "category": cat_obj,
                        "question": {"id": q_id, "en": "", "zh": q_zh},
                        "answer": {"id": a_id, "en": "", "zh": a_zh},
                    }
                )
        elif name == "11-Escalation":
            for row in rows[1:]:
                if not row:
                    continue
                scene, dept = _cell(row, 0), _cell(row, 1)
                if scene and dept:
                    escalation.append({"scene": scene, "department": dept})
        elif name == "12-QA标准":
            for row in rows[1:]:
                if not row:
                    continue
                item, standard = _cell(row, 0), _cell(row, 1)
                if item and standard:
                    qa_standards.append({"check": item, "standard": standard})

    return faq_items, escalation, qa_standards


def merge_with_legacy(v2: list[dict], legacy_path: Path) -> list[dict]:
    """Excel V2 first; keep unique legacy FAQ entries (with English) that are not near-duplicates."""
    merged: list[dict] = []
    seen: set[str] = set()

    for item in v2:
        q = item["question"]
        keys = {_norm(q.get("zh", "")), _norm(q.get("id", "")), _norm(q.get("en", ""))}
        keys.discard("")
        if any(k in seen for k in keys):
            continue
        seen.update(keys)
        merged.append(item)

    if legacy_path.exists():
        legacy = json.loads(legacy_path.read_text(encoding="utf-8"))
        for item in legacy:
            q = item.get("question") or {}
            keys = {_norm(q.get("zh", "")), _norm(q.get("id", "")), _norm(q.get("en", ""))}
            keys.discard("")
            if not keys or any(k in seen for k in keys):
                # Enrich EN on matching V2 row when possible
                for m in merged:
                    mq = m["question"]
                    mk = {_norm(mq.get("zh", "")), _norm(mq.get("id", ""))}
                    mk.discard("")
                    if mk & keys:
                        if not mq.get("en") and q.get("en"):
                            mq["en"] = q["en"]
                        ma = m["answer"]
                        la = item.get("answer") or {}
                        if not ma.get("en") and la.get("en"):
                            ma["en"] = la["en"]
                        break
                continue
            seen.update(keys)
            out = dict(item)
            out["source"] = out.get("source") or "legacy_faq"
            merged.append(out)

    for i, item in enumerate(merged, start=1):
        item["id"] = i
    return merged


def main() -> int:
    parser = argparse.ArgumentParser(description="Import PinGo CS KB xlsx")
    parser.add_argument("xlsx", nargs="?", default=str(DEFAULT_XLSX))
    parser.add_argument("--faq-out", default=str(KNOWLEDGE / "faq.json"))
    parser.add_argument("--kb-out", default=str(KNOWLEDGE / "kb_v2.json"))
    args = parser.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        print(f"xlsx not found: {xlsx}", file=sys.stderr)
        return 1

    KNOWLEDGE.mkdir(parents=True, exist_ok=True)
    v2_raw, escalation, qa_standards = load_xlsx(xlsx)

    # Snapshot of pure Excel rows (numbered)
    kb_v2 = []
    for i, item in enumerate(v2_raw, start=1):
        row = dict(item)
        row["id"] = i
        kb_v2.append(row)

    faq_path = Path(args.faq_out)
    # If faq already is a previous merge, prefer keeping a copy of original as faq_legacy once
    legacy_path = KNOWLEDGE / "faq_legacy.json"
    if not legacy_path.exists() and faq_path.exists():
        legacy_path.write_text(faq_path.read_text(encoding="utf-8"), encoding="utf-8")
        print(f"backed up previous faq → {legacy_path}")

    merged = merge_with_legacy(v2_raw, legacy_path if legacy_path.exists() else faq_path)

    Path(args.kb_out).write_text(
        json.dumps(kb_v2, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    faq_path.write_text(json.dumps(merged, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    (KNOWLEDGE / "escalation.json").write_text(
        json.dumps(escalation, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (KNOWLEDGE / "qa_standards.json").write_text(
        json.dumps(qa_standards, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(f"imported_xlsx_qa={len(kb_v2)}")
    print(f"faq_merged_total={len(merged)}")
    print(f"escalation={len(escalation)} qa_standards={len(qa_standards)}")
    print(f"wrote {args.kb_out}")
    print(f"wrote {args.faq_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
